from pathlib import Path
import os
import scrapy
from scrapy.settings import Settings
from scrapy.utils.log import configure_logging
from scrapy.crawler import CrawlerProcess
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, NamedStyle
from datetime import datetime
from decouple import config
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import smtplib
import re

from utils import get_specific_folder

currency_style = NamedStyle(name="currency_BRL")
currency_style.number_format = 'R$ #,##0.00'


class DiscosSpider(scrapy.Spider):
    name = "bot_ml"
    custom_settings = {
        "DEFAULT_REQUEST_HEADERS": {
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "pt",
        },
        # "ITEM_PIPELINES": {
        #     # "__main__.BotMlPipeline": 300,  # Pipeline declarado abaixo
        # },
        "DOWNLOADER_MIDDLEWARES": {
            "scrapy.downloadermiddlewares.useragent.UserAgentMiddleware": None,
            "scrapy_user_agents.middlewares.RandomUserAgentMiddleware": 500,
        },
    }

    def __init__(self, search_query, *args, **kwargs):
        super(DiscosSpider, self).__init__(*args, **kwargs)
        self.search_query = search_query
        self.start_urls_list_ml = "https://lista.mercadolivre.com.br"
        self.url_product = "https://produto.mercadolivre.com.br"


    def start_requests(self):
        if self.search_query:
            search_url = f"{self.start_urls_list_ml}/{self.search_query}#D[A:{self.search_query}]"
            yield scrapy.Request(url=search_url, callback=self.parse)
        else:
            self.logger.error("Nenhum termo de pesquisa fornecido. Use -a search_query='termo' para especificar.")



    def parse(self, response, **kwargs):
        for item in response.xpath("//li[@class='ui-search-layout__item shops__layout-item']"):
            title = item.xpath(".//h3[@class='poly-component__title-wrapper']//a[@class='poly-component__title']//text()").get()
            price = item.xpath(".//div[@class='poly-price__current']//span/span[@class='andes-money-amount__fraction']//text()").get()
            links = item.xpath(".//h3[@class='poly-component__title-wrapper']/a/@href").getall()
            link = links[0] if links else None

            processed_link = self.process_link(link)
            # title_exists = processed_link.split(f"{self.base}/", 1)[1]

            if processed_link:

                print({
                'titulo': title.strip() if title else None,
                'valor': price.strip() if price else None,  # Corrigido para verificar se price é válido
                'link': processed_link,
                'termo': self.search_query.lower().replace(" ", "_")
            })
                 
                yield {
                    'titulo': title.strip() if title else None,
                    'valor': price.strip() if f'R$ {price}' else None,
                    'link': processed_link,
                    'termo': self.search_query.lower().replace(" ", "_")
                }

        # pagination_links = response.xpath('//nav[@aria-label="Paginação"]//a/@href').getall()

        # for link in pagination_links:
        #     yield response.follow(link, self.parse)
        

    def process_link(self, link):
        """Processa o link, filtrando links que contêm 'MLB' e truncando até o '#'."""
        if link and 'MLB' in link:
            filtered_link = link.split('#')[0]
            return filtered_link
        return None
    

class BotMlPipeline:
    def __init__(self, csv_path):
        print("✅ Pipeline BotMlPipeline inicializada")
        self.results = []
        self.csv_path = csv_path
        self.date_file = datetime.now().strftime("%d-%m-%Y")

    # def process_item(self, item, spider):
    #     print(f"📌 Item processado no pipeline: {item}")
    #     self.results.append(item)
    #     return item

    def close_spider_teste(self, spider):
        print("🛑 Pipeline encerrando... Salvando Excel!")
        # df = pd.DataFrame(self.results)
        df = pd.read_csv(self.csv_path)
        termo = df.iloc[0]['termo'] if 'termo' in df.columns else ""
        # Crie um título seguro para a planilha
        safe_title = re.sub(r'[\\/?:*"[<>|]', '_', f"Resultados-{termo}-{self.date_file}")
        
        # Salve o DataFrame em um arquivo Excel
        output_file = f"{safe_title}.xlsx"
        df.to_excel(output_file, sheet_name=safe_title, index=False)
        
        spider.logger.info(f"Resultados salvos em {output_file}")

        # output_file = f"resultados-{self.date_file}.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = safe_title

        header_font = Font(bold=True, size=12)
        headers = [col.upper() for col in df.columns.tolist()] 
        ws.append(headers)
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for _, row in df.iterrows():
            current_row = ws.max_row + 1
            for col_num, (col_name, value) in enumerate(row.items(), start=1):
                cell = ws.cell(row=current_row, column=col_num)

                if col_name == 'link' and pd.notna(value):
                    cell.hyperlink = value
                    cell.font = Font(color="0000FF", underline="single")
                elif col_name == 'valor' and pd.notna(value):
                    cell.value = float(value)
                    cell.style = currency_style
                else:
                    cell.value = value

        # Ajusta a largura das colunas automaticamente
        # for col in ws.columns:
        #     max_length = max(len(str(cell.value) if cell.value else "") for cell in col)
        #     ws.column_dimensions[col[0].column_letter].width = max_length + 2

        file_path = os.path.join(get_specific_folder("output"), output_file)
        wb.save(file_path)
        spider.logger.info(f"Resultados salvos em {file_path}")
        # self.send_email(file_path)
        print(f"Arquivo Excel salvo em: {file_path}")



    def process_and_send_report(self):
        """Converte CSV para Excel e envia por e-mail"""
        df = pd.read_csv(self.csv_path)

        if df.empty:
            print("❌ Nenhum resultado encontrado. Nenhum e-mail será enviado.")
            return

        termo = df.iloc[0]['termo'] if 'termo' in df.columns else "pesquisa"
        safe_title = f"Resultados-{termo}-{self.date_file}".replace("/", "_")
        output_file = f"{safe_title}.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = safe_title

        header_font = Font(bold=True, size=12)
        headers = [col.upper() for col in df.columns.tolist()]
        ws.append(headers)

        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for _, row in df.iterrows():
            current_row = ws.max_row + 1
            for col_num, (col_name, value) in enumerate(row.items(), start=1):
                cell = ws.cell(row=current_row, column=col_num)

                if col_name == 'link' and pd.notna(value):
                    cell.hyperlink = value
                    cell.font = Font(color="0000FF", underline="single")
                elif col_name == 'valor' and pd.notna(value):
                    cell.value = float(value)
                    cell.style = currency_style
                else:
                    cell.value = value

        file_path = os.path.join(get_specific_folder("output"), output_file)
        wb.save(file_path)
        print(f"✅ Arquivo Excel salvo em: {file_path}")

        self.send_email(file_path)


    def send_email(self, file_path):
        smtp_server = 'smtp.gmail.com' # config("SMTP_SERVER")
        smtp_port = 587 #config("SMTP_PORT")
        smtp_user = 'bottestes872@gmail.com' #config("SMTP_USER")
        smtp_password = 'S4nb*2wN5ePtxdKnkTqfZ^tbKxL8JPUU' #config("SMTP_PASSWORD")
        
        recipient = "146188@aluno.sertao.ifrs.edu.br"
        subject = "📌 Resultados da Pesquisa Scrapy"
        body = "Segue em anexo os resultados da sua busca em formato Excel."

        msg = MIMEMultipart()
        msg.attach(MIMEText(body, "plain"))
        msg["From"] = smtp_user
        msg["To"] = recipient
        msg["Subject"] = subject
        msg.attach(MIMEText(body, "plain"))

        with open(file_path, "rb") as attachment:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(attachment.read())
            encoders.encode_base64(part)
            part.add_header("Content-Disposition", f"attachment; filename={os.path.basename(file_path)}")
            msg.attach(part)

        try:
            server = smtplib.SMTP(smtp_server, smtp_port)
            server.starttls() 
            server.login(smtp_user, smtp_password)
            server.sendmail(smtp_user, recipient, msg.as_string())
            server.quit()
            print("✅ E-mail enviado com sucesso!")
        except Exception as e:
            print(f"❌ Erro ao enviar e-mail: {e}")



configure_logging({"LOG_FORMAT": "%(levelname)s: %(message)s"})
# if __name__ == "__main__":
#     folder_path = get_specific_folder("resource")
#     date_str = datetime.now().strftime("%Y-%m-%d")
#     filename = f"resultado-{date_str}.csv"
#     file_path = os.path.join(folder_path, filename)
#     settings = Settings()
#     settings.set("FEEDS", {
#         file_path : {
#             "format": "csv",
#             "encoding": "utf-8-sig", 
#             },
#     })

#     process = CrawlerProcess(Settings())
#     process.crawl(DiscosSpider, search_query="Hareton Salvanini")
#     process.start()

def run_spiders():
    """Executa múltiplos spiders e gera o relatório final"""
    folder_path = get_specific_folder("resource")
    date_str = datetime.now().strftime("%Y-%m-%d")
    csv_path = os.path.join(folder_path, f"resultados-{date_str}.csv")

    settings = {
        "FEEDS": {
            csv_path: {
                "format": "csv",
                "encoding": "utf-8-sig",
                "overwrite": False,  # Para adicionar dados ao invés de sobrescrever
            },
        }
    }

    process = CrawlerProcess(settings)
    
    termos_de_busca = ["Hareton Salvanini", "Vinil Beatles", "CD Pink Floyd"]

    for termo in termos_de_busca:
        process.crawl(DiscosSpider, search_query=termo)

    process.start()  # Inicia o Scrapy e bloqueia até terminar todas as execuções

    # pipeline = BotMlPipeline(csv_path)
    # pipeline.process_and_send_report() 


if __name__ == "__main__":
    run_spiders()