import os
import re
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, NamedStyle
from utils import get_project_root, send_email


class BotMlPipeline:
    def __init__(self):
        print("✅ Pipeline BotMlPipeline inicializada")
        self.date_file = datetime.now().strftime("%d-%m-%Y_%H-%M")
        self.output_path = "output/"

    def get_latest_csv(self):
        """Encontra o arquivo CSV mais recente na pasta output/"""
        pattern = re.compile(r"resultados_(\d{2}-\d{2}-\d{4})_(\d{2}-\d{2})\.csv")
        files = []

        for filename in os.listdir(self.output_path):
            match = pattern.match(filename)
            if match:
                date_str, time_str = match.groups()
                full_datetime = datetime.strptime(f"{date_str} {time_str}", "%d-%m-%Y %H-%M")
                files.append((full_datetime, filename))

        if files:
            files.sort(reverse=True)  
            return os.path.join(self.output_path, files[0][1])
        return None

    def close_spider(self, spider):
        print("🛑 Pipeline encerrando... Salvando Excel!")

        latest_csv = self.get_latest_csv()
        if not latest_csv:
            print("⚠️ Nenhum arquivo CSV encontrado. Encerrando pipeline.")
            return

        print(f"📂 Último arquivo CSV encontrado: {latest_csv}")
        df = pd.read_csv(latest_csv)

        termo = df.iloc[0]['termo'] if 'termo' in df.columns else "resultados"
        safe_title = re.sub(r'[\\/?:*"[<>|]', '_', f"Resultados-{self.date_file}")

        output_file = f"{safe_title}.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = safe_title


        currency_style = NamedStyle(name="currency_BRL")
        currency_style.number_format = 'R$ #,##0.00'


        header_font = Font(bold=True, size=12)
        headers = [col.upper() for col in df.columns.tolist()]
        ws.append(headers)

        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")


        for _, row in df.iterrows():
            current_row = ws.max_row + 1
            for col_num, (col_name, value) in enumerate(row.items(), start=1):
                cell = ws.cell(row=current_row, column=col_num)

                if col_name == 'link' and pd.notna(value):
                    cell.hyperlink = value
                    cell.font = Font(color="0000FF", underline="single")
                elif col_name == 'valor' and pd.notna(value):
                    cleaned_value = re.sub(r"[^\d.]", "", str(value)) 
                    cell.value = cleaned_value
                    cell.style = currency_style
                else:
                    cell.value = value
        path_save_file = os.path.join(get_project_root(), "exel", output_file)
        wb.save(path_save_file)
        send_email(output_file, path_save_file)
        print(f"✅ Arquivo Excel salvo em: {path_save_file}")
        print(f"Resultados salvos em {path_save_file}")
        path_output_files = os.path.join(get_project_root(), "output")
        for filename in os.listdir(path_output_files):
            file_path = os.path.join(path_output_files, filename)
            if os.path.isfile(file_path):
                os.remove(file_path)
                print(f"Arquivo {filename} deletado.")
